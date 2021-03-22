import openpyxl as xl
from openpyxl.styles import Alignment

"""This file exports the results of a simulation to a spreadsheet"""
def export(path,fifo,fifo_inter,lru,lru_inter,job_list):
    """Params for sheet"""
    wb = xl.Workbook()
    
   
    
    worksheet = wb.active
    worksheet.title = 'My Simulation'
    worksheet.column_dimensions['A'].width = 20

    this_row = 1
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "FIFO Simulation"

    """Fifo Events"""
    for fifo_event in fifo:
        this_row = int(fifo_event.frame) + 1
        that_cell = worksheet.cell(row=this_row,column=1)
        that_cell.value = "Page Frame "+ str(fifo_event.frame)
        this_column = 2
        for event in fifo_event.event:
            if (event is None):
                event = "(empty)"
            this_cell = worksheet.cell(row = this_row, column = this_column)
            this_cell.value = event
            this_column = this_column +1
    
    this_row = this_row +2
    """Fifo Interrupts"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Interrupt"
    for inter in fifo_inter:
        for interrupt in inter:
            if(not(interrupt == "*")):
                interrupt = "-"
            this_cell = worksheet.cell(row = this_row, column = this_column)
            this_cell.value = interrupt
            this_column = this_column +1

    this_row = this_row +1
    """Fifo Page requests"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Page Requested"
    for job in job_list:
        this_cell = worksheet.cell(row = this_row, column = this_column)
        this_cell.value = job
        this_column = this_column +1

    this_row = this_row +1
    """Fifo Time snap"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Time Snapshot"
    for i in range(len(job_list)):
        this_cell = worksheet.cell(row = this_row, column = this_column)
        this_cell.value = i
        this_cell.alignment = Alignment(horizontal='left')
        this_column = this_column +1

    this_row = this_row + 3
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "LRU Simulation"
    
    this_column = 2
    
    """LRU Events"""
    for lru_event in lru:
        this_row = this_row + 1
        that_cell = worksheet.cell(row=this_row,column=1)
        that_cell.value = "Page Frame "+ str(lru_event.frame)
        this_column = 2
        for event in lru_event.event:
            if (event is None):
                event = "(empty)"
            this_cell = worksheet.cell(row = this_row, column = this_column)
            this_cell.value = event
            this_column = this_column +1

    this_row = this_row +2
    """LRU Interrupts"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Interrupt"
    for inter in lru_inter:
        for interrupt in inter:
            if(not(interrupt == "*")):
                interrupt = "-"
            this_cell = worksheet.cell(row = this_row, column = this_column)
            this_cell.value = interrupt
            this_column = this_column +1
        
    this_row = this_row +1
    
    """LRU Page requests"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Page Requested"
    for job in job_list:
        this_cell = worksheet.cell(row = this_row, column = this_column)
        this_cell.value = job
        this_column = this_column +1

    this_row = this_row +1
    """LRU Time snap"""
    this_column = 2
    that_cell = worksheet.cell(row=this_row,column=1)
    that_cell.value = "Time Snapshot"
    for i in range(len(job_list)):
        this_cell = worksheet.cell(row = this_row, column = this_column)
        this_cell.value = i
        this_cell.alignment = Alignment(horizontal='left')
        this_column = this_column +1

        
    wb.save(filename = path.name)

    #skip row for inters
    this_row = this_row + 1



